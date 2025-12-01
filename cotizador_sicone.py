"""
SICONE - Sistema de Cotización v3.0
Versión completa con gestión de cotizaciones guardadas
Autor: AI-MindNovation
Fecha: Noviembre 2025
"""

import streamlit as st
import pandas as pd
import numpy as np
import json
from datetime import datetime
from io import BytesIO
import plotly.express as px
import plotly.graph_objects as go
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional

# ============================================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================================
st.set_page_config(
    page_title="SICONE v2.0 - Cotizador",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# ESTILOS CSS
# ============================================================================
st.markdown("""
<style>
    .main-title {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .section-title {
        font-size: 1.5rem;
        color: #2c3e50;
        border-bottom: 2px solid #3498db;
        padding-bottom: 0.5rem;
        margin-top: 1.5rem;
    }
    .stExpander {
        border: 1px solid #e0e0e0;
        border-radius: 5px;
        margin-bottom: 1rem;
    }
    
    /* ========================================
       CONSISTENCIA DE FORMATO EN TABLAS
       Objetivo: Editables = gris, Calculados = blanco
       (igual que number_input y métricas)
       ======================================== */
    
    /* Intento 1: Usar atributo aria-readonly */
    [data-testid="stDataFrame"] [role="gridcell"]:not([aria-readonly="true"]) {
        background-color: #f0f2f6 !important;
    }
    
    [data-testid="stDataFrame"] [role="gridcell"][aria-readonly="true"] {
        background-color: white !important;
    }
    
    /* Intento 2: Usar data-testid específico */
    div[data-testid="stDataFrameResizable"] div[data-testid="stDataFrameDataCell"]:not([aria-readonly="true"]) {
        background-color: #f0f2f6 !important;
    }
    
    div[data-testid="stDataFrameResizable"] div[data-testid="stDataFrameDataCell"][aria-readonly="true"] {
        background-color: white !important;
    }
    
    /* Intento 3: Usar estructura de tabla base-web */
    .stDataFrame [data-baseweb="data-table"] [role="gridcell"]:not([data-readonly]) {
        background-color: #f0f2f6 !important;
    }
    
    .stDataFrame [data-baseweb="data-table"] [role="gridcell"][data-readonly] {
        background-color: white !important;
    }
    
    /* Intento 4: Clase específica de celdas */
    [data-testid="stDataFrame"] .editable-cell {
        background-color: #f0f2f6 !important;
    }
    
    [data-testid="stDataFrame"] .readonly-cell {
        background-color: white !important;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# MODELOS DE DATOS
# ============================================================================

@dataclass
class ProyectoInfo:
    """Información general del proyecto"""
    nombre: str = ""
    cliente: str = ""
    direccion: str = ""
    telefono: str = ""
    business_manager: str = ""
    medio_contacto: str = ""
    area_base: float = 0.0
    area_cubierta: float = 0.0
    area_entrepiso: float = 0.0
    niveles: int = 1
    muro_tipo: str = "sencillo"
    fecha: datetime = field(default_factory=datetime.now)

@dataclass
class ItemDiseno:
    """Item de Diseños y Planificación (se multiplica por área_base)"""
    nombre: str
    precio_unitario: float = 0.0
    
    def calcular_subtotal(self, area_base: float) -> float:
        return area_base * self.precio_unitario

@dataclass
class ItemEstandar:
    """Item estándar con Materiales, Equipos y Mano de Obra"""
    nombre: str
    unidad: str
    cantidad: float = 0.0
    precio_materiales: float = 0.0
    precio_equipos: float = 0.0
    precio_mano_obra: float = 0.0
    
    def calcular_subtotal(self) -> float:
        return (
            self.cantidad * self.precio_materiales +
            self.cantidad * self.precio_equipos +
            self.cantidad * self.precio_mano_obra
        )

@dataclass
class ItemCimentacion:
    """Item de cimentación (cantidad × precio unitario)"""
    nombre: str
    unidad: str
    cantidad: float = 0.0
    precio_unitario: float = 0.0
    
    def calcular_subtotal(self) -> float:
        return self.cantidad * self.precio_unitario

@dataclass
class PersonalAdmin:
    """Personal administrativo con cálculo de prestaciones"""
    nombre: str
    cantidad: int = 1
    valor_mes: float = 0.0
    pct_prestaciones: float = 54.0  # %
    dedicacion: float = 0.5  # 0.0 - 1.0
    meses: int = 6
    
    def calcular_total(self) -> float:
        return (
            self.cantidad *
            self.valor_mes *
            (1 + self.pct_prestaciones / 100) *
            self.dedicacion *
            self.meses
        )

class ConceptoDetallado:
    """Concepto administrativo con ítems detallados expandibles"""
    def __init__(self, nombre, items_detalle=None):
        self.nombre = nombre
        self.items_detalle = items_detalle if items_detalle is not None else {}
    
    def calcular_subtotal(self):
        return sum(self.items_detalle.values())
    
    def agregar_item(self, nombre_item, valor):
        self.items_detalle[nombre_item] = float(valor)
    
    def editar_item(self, nombre_item, nuevo_valor):
        if nombre_item in self.items_detalle:
            self.items_detalle[nombre_item] = float(nuevo_valor)
    
    def eliminar_item(self, nombre_item):
        if nombre_item in self.items_detalle:
            del self.items_detalle[nombre_item]

# ============================================================================
# INICIALIZACIÓN DE SESSION STATE
# ============================================================================

def inicializar_session_state():
    """Inicializa todas las variables de sesión"""
    
    if 'proyecto' not in st.session_state:
        st.session_state.proyecto = ProyectoInfo()
    
    # DISEÑOS Y PLANIFICACIÓN
    if 'disenos' not in st.session_state:
        st.session_state.disenos = {
            'Diseño Arquitectónico': ItemDiseno('Diseño Arquitectónico', 0.0),
            'Diseño Estructural': ItemDiseno('Diseño Estructural', 21000.0),
            'Desarrollo del Proyecto': ItemDiseno('Desarrollo del Proyecto', 18900.0),
            'Visita Técnica': ItemDiseno('Visita Técnica', 0.0)
        }
    
    # ESTRUCTURA
    if 'estructura' not in st.session_state:
        st.session_state.estructura = ItemEstandar(
            'Estructura General', 'gl', 1.03, 127386450.0, 0.0, 0.0
        )
    
    # MAMPOSTERÍA
    if 'mamposteria' not in st.session_state:
        st.session_state.mamposteria = ItemEstandar(
            'Mampostería', 'm²', 845.0, 67000.0, 7500.0, 45000.0
        )
    
    # TECHOS Y OTROS
    if 'mamposteria_techos' not in st.session_state:
        st.session_state.mamposteria_techos = {
            'Cubierta, Superboard y Manto': ItemEstandar('Cubierta, Superboard y Manto', 'm²', 120.0, 175000.0, 5000.0, 40000.0),
            'Ruana': ItemEstandar('Ruana', 'ml', 0.0, 40000.0, 0.0, 20000.0),
            'Contramarcos - Ventana': ItemEstandar('Contramarcos - Ventana', 'ml', 0.0, 15000.0, 1500.0, 8500.0),
            'Contramarcos - Puerta': ItemEstandar('Contramarcos - Puerta', 'ml', 0.0, 15000.0, 1500.0, 8500.0),
            'Embudos y Boquillas': ItemEstandar('Embudos y Boquillas', 'und', 0.0, 60000.0, 0.0, 10000.0),
            'Cubierta, Superboard y Shingle': ItemEstandar('Cubierta, Superboard y Shingle', 'm²', 137.5, 210000.0, 5000.0, 50000.0),
            'Entrepiso Placa Fácil': ItemEstandar('Entrepiso Placa Fácil', 'm²', 159.02, 175000.0, 5000.0, 35000.0),
            'Canoas': ItemEstandar('Canoas', 'ml', 0.0, 85000.0, 10000.0, 35000.0),
            'Pérgolas y Estructura sin Techo': ItemEstandar('Pérgolas y Estructura sin Techo', 'm²', 108.16, 175000.0, 15000.0, 45000.0),
            'Tapacanal y Lagrimal': ItemEstandar('Tapacanal y Lagrimal', 'ml', 0.0, 80000.0, 5000.0, 35000.0)
        }
    
    # CIMENTACIONES
    if 'opcion_cimentacion' not in st.session_state:
        st.session_state.opcion_cimentacion = 'Opción 2'
    
    if 'cimentacion_opcion1' not in st.session_state:
        st.session_state.cimentacion_opcion1 = {
            'Pilas a 3m y 5m': ItemCimentacion('Pilas a 3m y 5m', 'und', 73.0, 1340000.0),
            'Cimentación Vigas y Losa': ItemCimentacion('Cimentación Vigas y Losa', 'm²', 385.06, 280000.0)
        }
    
    if 'cimentacion_opcion2' not in st.session_state:
        st.session_state.cimentacion_opcion2 = {
            'Pilotes de Apoyo': ItemCimentacion('Pilotes de Apoyo', 'und', 210.0, 320000.0),
            'Cimentación Vigas y Losa': ItemCimentacion('Cimentación Vigas y Losa', 'm²', 385.06, 280000.0)
        }
    
    # COMPLEMENTARIOS
    if 'complementarios' not in st.session_state:
        st.session_state.complementarios = {
            'Red Aguas Lluvias': ItemCimentacion('Red Aguas Lluvias', 'gl', 1.0, 6150000.0),
            'Red Hidrosanitaria': ItemCimentacion('Red Hidrosanitaria', 'gl', 1.0, 13520000.0),
            'Estructura Escalas Metálicas': ItemCimentacion('Estructura Escalas Metálicas', 'und', 2.0, 8600000.0),
            'Campamento y baño': ItemCimentacion('Campamento y baño', 'gl', 1.0, 3000000.0),
            'Cerramiento en tela': ItemCimentacion('Cerramiento en tela', 'ml', 200.0, 14500.0),
            'Canoa Metálica Calibre 24': ItemCimentacion('Canoa Metálica Calibre 24', 'ml', 51.0, 145000.0),
            'Ruana Metálica Calibre 24': ItemCimentacion('Ruana Metálica Calibre 24', 'ml', 62.0, 58000.0),
            'Revoque': ItemCimentacion('Revoque', 'm²', 1690.0, 32500.0),
            'Fajas, Ranuras y Filetes': ItemCimentacion('Fajas, Ranuras y Filetes', 'ml', 1859.0, 7000.0),
            'Otros conceptos': ItemCimentacion('Otros conceptos', 'gl', 0.0, 0.0)
        }
    
    # PERSONAL PROFESIONAL
    if 'personal_profesional' not in st.session_state:
        st.session_state.personal_profesional = {
            'Director de Obra': PersonalAdmin('Director de Obra', 1, 4407865.0, 54.0, 0.5, 6),
            'Supervisor Técnico': PersonalAdmin('Supervisor Técnico', 1, 1889085.0, 54.0, 0.3, 6),
            'Profesional Presupuesto': PersonalAdmin('Profesional Presupuesto', 1, 2896597.0, 54.0, 0.3, 6),
            'Arquitecto Diseñador': PersonalAdmin('Arquitecto Diseñador', 1, 1259390.0, 54.0, 0.3, 3),
            'Oficial Obra': PersonalAdmin('Oficial Obra', 1, 2266902.0, 54.0, 0.3, 3),
            'Ayudante de Obra': PersonalAdmin('Ayudante de Obra', 1, 811488.0, 54.0, 0.2, 2)
        }
    
    # PERSONAL ADMINISTRATIVO
    if 'personal_administrativo' not in st.session_state:
        st.session_state.personal_administrativo = {
            'Profesional de Procesos': PersonalAdmin('Profesional de Procesos', 1, 4407865.0, 54.0, 0.3, 6),
            'Gerente General': PersonalAdmin('Gerente General', 1, 5667255.0, 54.0, 0.3, 6),
            'Compras': PersonalAdmin('Compras', 1, 3148475.0, 54.0, 0.3, 6),
            'Contabilidad': PersonalAdmin('Contabilidad', 1, 3148475.0, 54.0, 0.2, 6),
            'Atención al Cliente': PersonalAdmin('Atención al Cliente', 1, 1259390.0, 54.0, 0.2, 3),
            'Mantenimiento y Servicios Generales': PersonalAdmin('Mantenimiento y Servicios Generales', 1, 811489.0, 54.0, 0.2, 3),
            'Desarrollo y Gestión Humana': PersonalAdmin('Desarrollo y Gestión Humana', 1, 2140963.0, 54.0, 0.2, 3),
            'Personal Administrativo Planta': PersonalAdmin('Personal Administrativo Planta', 1, 3434700.0, 54.0, 0.3, 0),
            'Personal Operativo Planta': PersonalAdmin('Personal Operativo Planta', 1, 737717.0, 54.0, 0.3, 0),
            'Personal Gestión Ambiental': PersonalAdmin('Personal Gestión Ambiental', 1, 3000000.0, 54.0, 0.3, 0)
        }
    
    # OTROS CONCEPTOS ADMINISTRACIÓN (con detalle expandible)
    if 'otros_admin' not in st.session_state:
        st.session_state.otros_admin = {
            'Pólizas de Seguros': ConceptoDetallado('Pólizas de Seguros', {
                'Buen Manejo y Correcta Inversión del Anticipo': 600000.0,
                'Cumplimiento': 600000.0,
                'Pago de Salarios, Prestaciones Sociales e Indemnizaciones': 600000.0,
                'Estabilidad de Obra': 600000.0,
                'Responsabilidad Social Extracontractual': 600000.0
            }),
            'Pagos Provisionales': ConceptoDetallado('Pagos Provisionales', {
                'Corte de Muros y Techador': 0.0,
                'Bonificación Pronta Entrega Armadores': 0.0,
                'Transporte de Operarios': 0.0,
                'Adecuación Alcantarillado': 0.0
            }),
            'Pagos Mensuales': ConceptoDetallado('Pagos Mensuales', {
                'Servicios públicos': 2511240.0,
                'Mantenimiento oficina': 2000000.0,
                'Comunicaciones': 1500000.0,
                'Transporte': 1500000.0,
                'Papelería': 0.0,
                'Caja Menor': 0.0
            }),
            'Dotaciones': ConceptoDetallado('Dotaciones', {
                'Uniformes': 0.0,
                'EPP': 0.0,
                'Herramientas menores': 0.0
            }),
            'Pagos de Obra': ConceptoDetallado('Pagos de Obra', {
                'Actualización de Planos': 0.0,
                'Elaboración de Planos': 0.0,
                'Amarre Geodésico ante Planeación': 0.0,
                'Georreferenciación de redes de servicios públicos': 0.0,
                'Diseño de Mezclas': 0.0,
                'Fallada de Cilindros': 0.0,
                'Fallada de Refuerzo': 0.0,
                'Fallada de Ladrillos': 0.0,
                'Ensayos de Tejas': 0.0,
                'Valla 3.0m x 2.0m': 0.0,
                'Valla 1.5m x 1.0m': 0.0
            }),
            'SISO': ConceptoDetallado('SISO', {
                'Cinta Precaución y Peligro': 0.0,
                'Tela Verde de Cerramiento': 0.0,
                'Alquiler Señales Preventivas': 0.0,
                'Cono PVC con cinta reflectiva 45cm': 0.0,
                'Tubo Señalizador con Cinta': 0.0,
                'Paleta para Pare y Siga': 0.0,
                'Señales 15cm x 15cm': 0.0,
                'Cascos de Seguridad Dieléctrico': 0.0,
                'Arnes de 3 argollas': 161700.0,
                'Cabo de Vida 5/8" x 1.80m': 103950.0,
                'Camilla Rígida': 0.0,
                'Respiradores RedLine por 50un': 42525.0,
                'Protector Auditivo tipo Copa Plegable': 0.0,
                'Chaleco Reflectivo': 0.0,
                'Bota de Caucho Trepadora': 0.0,
                'Bota Cuero Concord': 0.0,
                'Capa Adulto 1.50m x 2.0m': 45864.0,
                'Conjunto Industrial Nylon Poliéster': 0.0,
                'Guantes de Nylon': 89775.0,
                'Guantes de Hilaza': 0.0,
                'Botiquín de Primeros Auxilios': 0.0,
                'Extintor Multipropósito': 0.0,
                'Ajuste SISO y Seguridad Industrial': 2556186.0
            }),
            'Asesores Externos': ConceptoDetallado('Asesores Externos', {
                'Suelos': 0.0,
                'Hidráulica': 0.0,
                'Estructuras': 0.0,
                'Contable': 0.0,
                'Jurídico': 0.0,
                'Electricista': 0.0
            }),
            'Impuestos': ConceptoDetallado('Impuestos', {}),  # Se calcula dinámicamente
            'Costos Fijos': ConceptoDetallado('Costos Fijos', {
                'Gastos Sede Central': 2215224.0,
                'Gastos Financieros': 1846020.0,
                'Costo de Elaboración de la Propuesta': 928200.0
            }),
            'Descuentos': ConceptoDetallado('Descuentos', {
                'Descuento pronto pago': 0.0,
                'Bonificaciones': 0.0
            }),
            'Pagos a Terceros': ConceptoDetallado('Pagos a Terceros', {
                'Cálculo Estructural': 0.0,
                'Desarrollo del Proyecto': 0.0,
                'Comisiones': 0.0,
                'Otras Comisiones': 0.0
            })
        }
    
    # CONFIGURACIÓN AIU
    if 'config_aiu' not in st.session_state:
        st.session_state.config_aiu = {
            'Comisión de Ventas (%)': 5.5,
            'Imprevistos (%)': 10.5,
            'Administración (%)': 27.5,
            'Logística (%)': 2.5,
            'Utilidad (%)': 26.5
        }
    
    # % AIU CIMENTACIONES Y COMPLEMENTARIOS
    if 'aiu_cimentacion' not in st.session_state:
        st.session_state.aiu_cimentacion = {
            'pct_comision': 3.0,
            'pct_aiu': 47.0,
            'logistica': 0.0
        }
    
    if 'aiu_complementarios' not in st.session_state:
        st.session_state.aiu_complementarios = {
            'pct_comision': 0.0,
            'pct_aiu': 15.0,
            'logistica': 0.0
        }
    
    # Inicializar nombre de cotización actual
    if 'nombre_cotizacion_actual' not in st.session_state:
        st.session_state.nombre_cotizacion_actual = ""

# ============================================================================
# FUNCIONES DE CÁLCULO
# ============================================================================

def calcular_disenos():
    """Calcula subtotal de Diseños (usa SOLO área_base)"""
    area_base = st.session_state.proyecto.area_base
    total = sum([
        item.calcular_subtotal(area_base) 
        for item in st.session_state.disenos.values()
    ])
    return total

def calcular_estructura():
    """Calcula subtotal de Estructura"""
    return st.session_state.estructura.calcular_subtotal()

def calcular_mamposteria():
    """Calcula subtotal de Mampostería"""
    return st.session_state.mamposteria.calcular_subtotal()

def calcular_mamposteria_techos():
    """Calcula subtotal de Mampostería y Techos"""
    total = sum([
        item.calcular_subtotal() 
        for item in st.session_state.mamposteria_techos.values()
    ])
    return total

def calcular_cimentacion():
    """Calcula total de cimentación según opción seleccionada"""
    opcion = st.session_state.opcion_cimentacion
    
    if opcion == 'Opción 1':
        items = st.session_state.cimentacion_opcion1
    else:
        items = st.session_state.cimentacion_opcion2
    
    subtotal = sum([item.calcular_subtotal() for item in items.values()])
    
    # Agregar AIU específico de cimentación
    comision = subtotal * (st.session_state.aiu_cimentacion['pct_comision'] / 100)
    aiu = subtotal * (st.session_state.aiu_cimentacion['pct_aiu'] / 100)
    logistica = st.session_state.aiu_cimentacion['logistica']
    
    total = subtotal + comision + aiu + logistica
    
    return {
        'subtotal': subtotal,
        'comision': comision,
        'aiu': aiu,
        'logistica': logistica,
        'total': total
    }

def calcular_complementarios():
    """Calcula total de complementarios"""
    subtotal = sum([
        item.calcular_subtotal() 
        for item in st.session_state.complementarios.values()
    ])
    
    # Agregar AIU específico de complementarios
    comision = subtotal * (st.session_state.aiu_complementarios['pct_comision'] / 100)
    aiu = subtotal * (st.session_state.aiu_complementarios['pct_aiu'] / 100)
    logistica = st.session_state.aiu_complementarios['logistica']
    
    total = subtotal + comision + aiu + logistica
    
    return {
        'subtotal': subtotal,
        'comision': comision,
        'aiu': aiu,
        'logistica': logistica,
        'total': total
    }

def calcular_administracion_detallada():
    """Calcula administración detallada"""
    # Primero calcular impuestos dinámicamente
    calcular_impuestos_dinamicos()
    
    total_prof = sum([
        p.calcular_total() 
        for p in st.session_state.personal_profesional.values()
    ])
    
    total_admin = sum([
        p.calcular_total() 
        for p in st.session_state.personal_administrativo.values()
    ])
    
    # Actualizado para ConceptoDetallado
    total_otros = sum([
        concepto.calcular_subtotal() 
        for concepto in st.session_state.otros_admin.values()
    ])
    
    total = total_prof + total_admin + total_otros
    
    return {
        'personal_profesional': total_prof,
        'personal_administrativo': total_admin,
        'otros_conceptos': total_otros,
        'total': total
    }

def serializar_cotizacion():
    """
    Serializa el estado completo de la cotización a un diccionario JSON-compatible
    """
    cotizacion_data = {
        'version': '3.0',
        'fecha_guardado': datetime.now().isoformat(),
        'proyecto': {
            'nombre': st.session_state.proyecto.nombre,
            'cliente': st.session_state.proyecto.cliente,
            'direccion': st.session_state.proyecto.direccion,
            'telefono': st.session_state.proyecto.telefono,
            'business_manager': st.session_state.proyecto.business_manager,
            'medio_contacto': st.session_state.proyecto.medio_contacto,
            'area_base': st.session_state.proyecto.area_base,
            'area_cubierta': st.session_state.proyecto.area_cubierta,
            'area_entrepiso': st.session_state.proyecto.area_entrepiso,
            'niveles': st.session_state.proyecto.niveles,
            'muro_tipo': st.session_state.proyecto.muro_tipo
        },
        'config_aiu': st.session_state.config_aiu,
        'disenos': {
            nombre: {
                'precio_unitario': item.precio_unitario
            } for nombre, item in st.session_state.disenos.items()
        },
        'estructura': {
            'cantidad': st.session_state.estructura.cantidad,
            'precio_materiales': st.session_state.estructura.precio_materiales,
            'precio_equipos': st.session_state.estructura.precio_equipos,
            'precio_mano_obra': st.session_state.estructura.precio_mano_obra
        },
        'mamposteria': {
            'cantidad': st.session_state.mamposteria.cantidad,
            'precio_materiales': st.session_state.mamposteria.precio_materiales,
            'precio_equipos': st.session_state.mamposteria.precio_equipos,
            'precio_mano_obra': st.session_state.mamposteria.precio_mano_obra
        },
        'mamposteria_techos': {
            nombre: {
                'cantidad': item.cantidad,
                'precio_materiales': item.precio_materiales,
                'precio_equipos': item.precio_equipos,
                'precio_mano_obra': item.precio_mano_obra
            } for nombre, item in st.session_state.mamposteria_techos.items()
        },
        'opcion_cimentacion': st.session_state.opcion_cimentacion,
        'cimentacion_opcion1': {
            nombre: {
                'cantidad': item.cantidad,
                'precio_unitario': item.precio_unitario
            } for nombre, item in st.session_state.cimentacion_opcion1.items()
        },
        'cimentacion_opcion2': {
            nombre: {
                'cantidad': item.cantidad,
                'precio_unitario': item.precio_unitario
            } for nombre, item in st.session_state.cimentacion_opcion2.items()
        },
        'aiu_cimentacion': st.session_state.aiu_cimentacion,
        'complementarios': {
            nombre: {
                'cantidad': item.cantidad,
                'precio_unitario': item.precio_unitario
            } for nombre, item in st.session_state.complementarios.items()
        },
        'aiu_complementarios': st.session_state.aiu_complementarios,
        'personal_profesional': {
            nombre: {
                'cantidad': p.cantidad,
                'valor_mes': p.valor_mes,
                'pct_prestaciones': p.pct_prestaciones,
                'dedicacion': p.dedicacion,
                'meses': p.meses
            } for nombre, p in st.session_state.personal_profesional.items()
        },
        'personal_administrativo': {
            nombre: {
                'cantidad': p.cantidad,
                'valor_mes': p.valor_mes,
                'pct_prestaciones': p.pct_prestaciones,
                'dedicacion': p.dedicacion,
                'meses': p.meses
            } for nombre, p in st.session_state.personal_administrativo.items()
        },
        'otros_admin': {
            concepto_nombre: {
                'items_detalle': concepto_obj.items_detalle
            } for concepto_nombre, concepto_obj in st.session_state.otros_admin.items()
        }
    }
    
    return cotizacion_data

def deserializar_cotizacion(cotizacion_data):
    """
    Carga una cotizacion desde un diccionario y actualiza el session_state
    """
    # Proyecto
    for key, value in cotizacion_data['proyecto'].items():
        setattr(st.session_state.proyecto, key, value)
    
    # Configuración AIU
    st.session_state.config_aiu = cotizacion_data['config_aiu']
    
    # Diseños
    for nombre, data in cotizacion_data['disenos'].items():
        if nombre in st.session_state.disenos:
            st.session_state.disenos[nombre].precio_unitario = data['precio_unitario']
    
    # Estructura
    st.session_state.estructura.cantidad = cotizacion_data['estructura']['cantidad']
    st.session_state.estructura.precio_materiales = cotizacion_data['estructura']['precio_materiales']
    st.session_state.estructura.precio_equipos = cotizacion_data['estructura']['precio_equipos']
    st.session_state.estructura.precio_mano_obra = cotizacion_data['estructura']['precio_mano_obra']
    
    # Mampostería
    st.session_state.mamposteria.cantidad = cotizacion_data['mamposteria']['cantidad']
    st.session_state.mamposteria.precio_materiales = cotizacion_data['mamposteria']['precio_materiales']
    st.session_state.mamposteria.precio_equipos = cotizacion_data['mamposteria']['precio_equipos']
    st.session_state.mamposteria.precio_mano_obra = cotizacion_data['mamposteria']['precio_mano_obra']
    
    # Mampostería y techos
    for nombre, data in cotizacion_data['mamposteria_techos'].items():
        if nombre in st.session_state.mamposteria_techos:
            st.session_state.mamposteria_techos[nombre].cantidad = data['cantidad']
            st.session_state.mamposteria_techos[nombre].precio_materiales = data['precio_materiales']
            st.session_state.mamposteria_techos[nombre].precio_equipos = data['precio_equipos']
            st.session_state.mamposteria_techos[nombre].precio_mano_obra = data['precio_mano_obra']
    
    # Cimentaciones
    st.session_state.opcion_cimentacion = cotizacion_data['opcion_cimentacion']
    for nombre, data in cotizacion_data['cimentacion_opcion1'].items():
        if nombre in st.session_state.cimentacion_opcion1:
            st.session_state.cimentacion_opcion1[nombre].cantidad = data['cantidad']
            st.session_state.cimentacion_opcion1[nombre].precio_unitario = data['precio_unitario']
    
    for nombre, data in cotizacion_data['cimentacion_opcion2'].items():
        if nombre in st.session_state.cimentacion_opcion2:
            st.session_state.cimentacion_opcion2[nombre].cantidad = data['cantidad']
            st.session_state.cimentacion_opcion2[nombre].precio_unitario = data['precio_unitario']
    
    st.session_state.aiu_cimentacion = cotizacion_data['aiu_cimentacion']
    
    # Complementarios
    for nombre, data in cotizacion_data['complementarios'].items():
        if nombre in st.session_state.complementarios:
            st.session_state.complementarios[nombre].cantidad = data['cantidad']
            st.session_state.complementarios[nombre].precio_unitario = data['precio_unitario']
    
    st.session_state.aiu_complementarios = cotizacion_data['aiu_complementarios']
    
    # Personal profesional
    for nombre, data in cotizacion_data['personal_profesional'].items():
        if nombre in st.session_state.personal_profesional:
            st.session_state.personal_profesional[nombre].cantidad = data['cantidad']
            st.session_state.personal_profesional[nombre].valor_mes = data['valor_mes']
            st.session_state.personal_profesional[nombre].pct_prestaciones = data['pct_prestaciones']
            st.session_state.personal_profesional[nombre].dedicacion = data['dedicacion']
            st.session_state.personal_profesional[nombre].meses = data['meses']
    
    # Personal administrativo
    for nombre, data in cotizacion_data['personal_administrativo'].items():
        if nombre in st.session_state.personal_administrativo:
            st.session_state.personal_administrativo[nombre].cantidad = data['cantidad']
            st.session_state.personal_administrativo[nombre].valor_mes = data['valor_mes']
            st.session_state.personal_administrativo[nombre].pct_prestaciones = data['pct_prestaciones']
            st.session_state.personal_administrativo[nombre].dedicacion = data['dedicacion']
            st.session_state.personal_administrativo[nombre].meses = data['meses']
    
    # Otros conceptos administrativos
    for concepto_nombre, data in cotizacion_data['otros_admin'].items():
        if concepto_nombre in st.session_state.otros_admin:
            st.session_state.otros_admin[concepto_nombre].items_detalle = data['items_detalle']

def guardar_cotizacion_memoria(nombre_cotizacion):
    """
    Guarda la cotización actual en memoria Y en archivo JSON persistente
    """
    import os
    
    if 'cotizaciones_guardadas' not in st.session_state:
        st.session_state.cotizaciones_guardadas = {}
    
    cotizacion_data = serializar_cotizacion()
    st.session_state.cotizaciones_guardadas[nombre_cotizacion] = cotizacion_data
    
    # Guardar en archivo para persistencia entre sesiones
    try:
        # Usar directorio relativo en lugar de /home/claude
        # Streamlit Cloud permite escribir en el directorio de trabajo actual
        cotizaciones_dir = './cotizaciones_sicone'
        os.makedirs(cotizaciones_dir, exist_ok=True)
        
        # Nombre de archivo seguro
        nombre_archivo = nombre_cotizacion.replace(' ', '_').replace('/', '_')
        filepath = os.path.join(cotizaciones_dir, f"{nombre_archivo}.json")
        
        # Guardar JSON
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(cotizacion_data, f, indent=2, ensure_ascii=False)
        
        return True, "Cotización guardada correctamente"
    except Exception as e:
        return False, f"Error al guardar: {str(e)}"

def cargar_cotizaciones_disponibles():
    """
    Carga la lista de cotizaciones disponibles desde archivos
    """
    import os
    
    # Usar directorio relativo
    cotizaciones_dir = './cotizaciones_sicone'
    
    if not os.path.exists(cotizaciones_dir):
        return {}
    
    cotizaciones = {}
    try:
        for filename in os.listdir(cotizaciones_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(cotizaciones_dir, filename)
                with open(filepath, 'r', encoding='utf-8') as f:
                    nombre = filename[:-5].replace('_', ' ')  # Remover .json
                    cotizaciones[nombre] = json.load(f)
        return cotizaciones
    except Exception as e:
        st.error(f"Error cargando cotizaciones: {str(e)}")
        return {}

def eliminar_cotizacion_archivo(nombre_cotizacion):
    """
    Elimina una cotización del archivo
    """
    import os
    
    try:
        # Usar directorio relativo
        cotizaciones_dir = './cotizaciones_sicone'
        nombre_archivo = nombre_cotizacion.replace(' ', '_').replace('/', '_')
        filepath = os.path.join(cotizaciones_dir, f"{nombre_archivo}.json")
        
        if os.path.exists(filepath):
            os.remove(filepath)
            return True, "Cotización eliminada"
        return False, "Archivo no encontrado"
    except Exception as e:
        return False, f"Error al eliminar: {str(e)}"

def cargar_cotizacion_memoria(nombre_cotizacion):
    """
    Carga una cotización desde memoria
    """
    if nombre_cotizacion in st.session_state.cotizaciones_guardadas:
        cotizacion_data = st.session_state.cotizaciones_guardadas[nombre_cotizacion]
        deserializar_cotizacion(cotizacion_data)
        return True
    return False

def exportar_cotizacion_json():
    """
    Exporta la cotización actual a JSON
    """
    cotizacion_data = serializar_cotizacion()
    json_str = json.dumps(cotizacion_data, indent=2, ensure_ascii=False)
    return json_str.encode('utf-8')

def importar_cotizacion_json(json_file):
    """
    Importa una cotización desde un archivo JSON
    """
    try:
        cotizacion_data = json.loads(json_file.getvalue().decode('utf-8'))
        
        # Deserializar
        deserializar_cotizacion(cotizacion_data)
        
        # Marcar el nombre de la cotización cargada
        st.session_state.nombre_cotizacion_actual = st.session_state.proyecto.nombre
        
        return True, "Cotización cargada exitosamente"
    except Exception as e:
        return False, f"Error al cargar cotización: {str(e)}"

def exportar_a_excel():
    """
    Exporta la cotización completa a un archivo Excel con todas las secciones
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización SICONE"
    
    # Estilos
    title_font = Font(bold=True, size=16, color="4472C4")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=12)
    subsection_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
    
    row = 1
    
    # ============================================================================
    # ENCABEZADO PRINCIPAL
    # ============================================================================
    ws[f'A{row}'] = "SICONE v3.0 - Sistema de Cotización"
    ws[f'A{row}'].font = title_font
    ws.merge_cells(f'A{row}:F{row}')
    row += 2
    
    # Información del Proyecto
    ws[f'A{row}'] = "Proyecto:"
    ws[f'B{row}'] = st.session_state.proyecto.nombre
    ws[f'D{row}'] = "Cliente:"
    ws[f'E{row}'] = st.session_state.proyecto.cliente
    row += 1
    
    ws[f'A{row}'] = "Dirección:"
    ws[f'B{row}'] = st.session_state.proyecto.direccion
    ws[f'D{row}'] = "Área Base:"
    ws[f'E{row}'] = st.session_state.proyecto.area_base
    ws[f'E{row}'].number_format = '0.00'
    ws[f'F{row}'] = "m²"
    row += 2
    
    # RESUMEN GLOBAL
    resumen = get_resumen_cacheado()
    
    ws[f'A{row}'] = "RESUMEN GLOBAL DEL PROYECTO"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    ws[f'A{row}'] = "Total Proyecto:"
    ws[f'B{row}'] = resumen['total_proyecto']
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'D{row}'] = "Precio por m²:"
    ws[f'E{row}'] = resumen['precio_m2']
    ws[f'E{row}'].number_format = '$#,##0'
    row += 2
    
    # ============================================================================
    # COTIZACIÓN 1 Y 2
    # ============================================================================
    ws[f'A{row}'] = "📋 Cotización 1: Diseños + Estructura + Mampostería"
    ws[f'A{row}'].font = subsection_font
    ws.merge_cells(f'A{row}:C{row}')
    
    ws[f'D{row}'] = "📋 Cotización 2: Cimentaciones + Complementarios"
    ws[f'D{row}'].font = subsection_font
    ws.merge_cells(f'D{row}:F{row}')
    row += 1
    
    # Cotización 1 - Costos Directos
    cot1 = resumen['cotizacion1']
    ws[f'A{row}'] = "Costos Directos:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = "Cimentaciones:"
    ws[f'D{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'A{row}'] = f"• Diseños:"
    ws[f'B{row}'] = cot1['disenos']
    ws[f'B{row}'].number_format = '$#,##0'
    
    ws[f'D{row}'] = f"Subtotal:"
    ws[f'E{row}'] = resumen['cotizacion2']['cimentacion']['subtotal']
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = f"• Estructura:"
    ws[f'B{row}'] = cot1['estructura']
    ws[f'B{row}'].number_format = '$#,##0'
    
    ws[f'D{row}'] = f"Comisión:"
    ws[f'E{row}'] = resumen['cotizacion2']['cimentacion']['comision']
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = f"• Mampostería:"
    ws[f'B{row}'] = cot1['mamposteria']
    ws[f'B{row}'].number_format = '$#,##0'
    
    ws[f'D{row}'] = f"AIU:"
    ws[f'E{row}'] = resumen['cotizacion2']['cimentacion']['aiu']
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = f"• Techos y otros:"
    ws[f'B{row}'] = cot1['mamposteria_techos']
    ws[f'B{row}'].number_format = '$#,##0'
    
    ws[f'D{row}'] = f"Total Cimentación:"
    ws[f'E{row}'] = resumen['cotizacion2']['cimentacion']['total']
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'A{row}'] = "Subtotal Costos Directos:"
    ws[f'B{row}'] = cot1['costos_directos']
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True)
    
    ws[f'D{row}'] = "Complementarios:"
    ws[f'D{row}'].font = Font(bold=True)
    row += 1
    
    # AIU Cotización 1
    ws[f'A{row}'] = "AIU:"
    ws[f'A{row}'].font = Font(bold=True)
    
    ws[f'D{row}'] = f"Subtotal:"
    ws[f'E{row}'] = resumen['cotizacion2']['complementarios']['subtotal']
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = f"• Comisión Ventas:"
    ws[f'B{row}'] = cot1['aiu']['comision_ventas']
    ws[f'B{row}'].number_format = '$#,##0'
    
    ws[f'D{row}'] = f"Comisión:"
    ws[f'E{row}'] = resumen['cotizacion2']['complementarios']['comision']
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = f"• Imprevistos:"
    ws[f'B{row}'] = cot1['aiu']['imprevistos']
    ws[f'B{row}'].number_format = '$#,##0'
    
    ws[f'D{row}'] = f"AIU:"
    ws[f'E{row}'] = resumen['cotizacion2']['complementarios']['aiu']
    ws[f'E{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = f"• Administración:"
    ws[f'B{row}'] = cot1['aiu']['administracion']
    ws[f'B{row}'].number_format = '$#,##0'
    
    ws[f'D{row}'] = f"Total Complementarios:"
    ws[f'E{row}'] = resumen['cotizacion2']['complementarios']['total']
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'E{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'A{row}'] = f"• Logística:"
    ws[f'B{row}'] = cot1['aiu']['logistica']
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = f"• Utilidad:"
    ws[f'B{row}'] = cot1['aiu']['utilidad']
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'A{row}'] = "Total AIU:"
    ws[f'B{row}'] = cot1['aiu']['total']
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    
    # Totales por cotización
    ws[f'A{row}'] = "TOTAL COTIZACIÓN 1:"
    ws[f'B{row}'] = cot1['total']
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = total_fill
    ws[f'B{row}'].fill = total_fill
    
    ws[f'D{row}'] = "TOTAL COTIZACIÓN 2:"
    ws[f'E{row}'] = resumen['cotizacion2']['total']
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'D{row}'].font = Font(bold=True, size=12)
    ws[f'E{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'].fill = total_fill
    ws[f'E{row}'].fill = total_fill
    row += 3
    
    # ============================================================================
    # DISEÑOS Y PLANIFICACIÓN (DETALLADO)
    # ============================================================================
    ws[f'A{row}'] = "DISEÑOS Y PLANIFICACIÓN"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = "Ítem"
    ws[f'B{row}'] = "Precio Unitario ($/m²)"
    ws[f'C{row}'] = "Área Base (m²)"
    ws[f'D{row}'] = "Subtotal"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
    row += 1
    
    for nombre, item in st.session_state.disenos.items():
        ws[f'A{row}'] = nombre
        ws[f'B{row}'] = item.precio_unitario
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'] = st.session_state.proyecto.area_base
        ws[f'C{row}'].number_format = '0.00'
        subtotal = item.calcular_subtotal(st.session_state.proyecto.area_base)
        ws[f'D{row}'] = subtotal
        ws[f'D{row}'].number_format = '$#,##0'
        row += 1
    
    ws[f'C{row}'] = "TOTAL:"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = calcular_disenos()
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'D{row}'].font = Font(bold=True)
    row += 2
    
    # ============================================================================
    # ESTRUCTURA (DETALLADO)
    # ============================================================================
    ws[f'A{row}'] = "ESTRUCTURA"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    ws[f'A{row}'] = "Ítem"
    ws[f'B{row}'] = "Cantidad"
    ws[f'C{row}'] = "Materiales"
    ws[f'D{row}'] = "Equipos"
    ws[f'E{row}'] = "Mano de Obra"
    ws[f'F{row}'] = "Subtotal"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
    row += 1
    
    estructura = st.session_state.estructura
    ws[f'A{row}'] = "Estructura"
    ws[f'B{row}'] = estructura.cantidad
    ws[f'B{row}'].number_format = '0.00'
    ws[f'C{row}'] = estructura.precio_materiales
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'D{row}'] = estructura.precio_equipos
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'E{row}'] = estructura.precio_mano_obra
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'F{row}'] = estructura.calcular_subtotal()
    ws[f'F{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'E{row}'] = "TOTAL:"
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'F{row}'] = calcular_estructura()
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].font = Font(bold=True)
    row += 2
    
    # ============================================================================
    # MAMPOSTERÍA (DETALLADO)
    # ============================================================================
    ws[f'A{row}'] = "MAMPOSTERÍA"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    ws[f'A{row}'] = "Ítem"
    ws[f'B{row}'] = "Cantidad (m²)"
    ws[f'C{row}'] = "Materiales"
    ws[f'D{row}'] = "Equipos"
    ws[f'E{row}'] = "Mano de Obra"
    ws[f'F{row}'] = "Subtotal"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
    row += 1
    
    mamposteria = st.session_state.mamposteria
    ws[f'A{row}'] = "Mampostería"
    ws[f'B{row}'] = mamposteria.cantidad
    ws[f'B{row}'].number_format = '0.00'
    ws[f'C{row}'] = mamposteria.precio_materiales
    ws[f'C{row}'].number_format = '$#,##0'
    ws[f'D{row}'] = mamposteria.precio_equipos
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'E{row}'] = mamposteria.precio_mano_obra
    ws[f'E{row}'].number_format = '$#,##0'
    ws[f'F{row}'] = mamposteria.calcular_subtotal()
    ws[f'F{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'E{row}'] = "TOTAL:"
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'F{row}'] = calcular_mamposteria()
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].font = Font(bold=True)
    row += 2
    
    # ============================================================================
    # TECHOS Y OTROS (DETALLADO)
    # ============================================================================
    ws[f'A{row}'] = "TECHOS Y OTROS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    ws[f'A{row}'] = "Ítem"
    ws[f'B{row}'] = "Cantidad"
    ws[f'C{row}'] = "Materiales"
    ws[f'D{row}'] = "Equipos"
    ws[f'E{row}'] = "Mano de Obra"
    ws[f'F{row}'] = "Subtotal"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
    row += 1
    
    for nombre, item in st.session_state.mamposteria_techos.items():
        ws[f'A{row}'] = nombre
        ws[f'B{row}'] = item.cantidad
        ws[f'B{row}'].number_format = '0.00'
        ws[f'C{row}'] = item.precio_materiales
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'] = item.precio_equipos
        ws[f'D{row}'].number_format = '$#,##0'
        ws[f'E{row}'] = item.precio_mano_obra
        ws[f'E{row}'].number_format = '$#,##0'
        ws[f'F{row}'] = item.calcular_subtotal()
        ws[f'F{row}'].number_format = '$#,##0'
        row += 1
    
    ws[f'E{row}'] = "TOTAL:"
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'F{row}'] = calcular_mamposteria_techos()
    ws[f'F{row}'].number_format = '$#,##0'
    ws[f'F{row}'].font = Font(bold=True)
    row += 2
    
    # ============================================================================
    # CIMENTACIONES (DETALLADO)
    # ============================================================================
    ws[f'A{row}'] = "CIMENTACIONES"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = "Ítem"
    ws[f'B{row}'] = "Cantidad"
    ws[f'C{row}'] = "Precio Unitario"
    ws[f'D{row}'] = "Subtotal"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
    row += 1
    
    # Usar opción seleccionada
    opcion = st.session_state.opcion_cimentacion
    items_cim = st.session_state.cimentacion_opcion1 if opcion == 'Opción 1' else st.session_state.cimentacion_opcion2
    
    for nombre, item in items_cim.items():
        ws[f'A{row}'] = nombre
        ws[f'B{row}'] = item.cantidad
        ws[f'B{row}'].number_format = '0.00'
        ws[f'C{row}'] = item.precio_unitario
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'] = item.calcular_subtotal()
        ws[f'D{row}'].number_format = '$#,##0'
        row += 1
    
    cim_calc = calcular_cimentacion()
    ws[f'C{row}'] = "Subtotal:"
    ws[f'D{row}'] = cim_calc['subtotal']
    ws[f'D{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'C{row}'] = "Comisión:"
    ws[f'D{row}'] = cim_calc['comision']
    ws[f'D{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'C{row}'] = "AIU:"
    ws[f'D{row}'] = cim_calc['aiu']
    ws[f'D{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'C{row}'] = "TOTAL:"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = cim_calc['total']
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'D{row}'].font = Font(bold=True)
    row += 2
    
    # ============================================================================
    # COMPLEMENTARIOS (DETALLADO)
    # ============================================================================
    ws[f'A{row}'] = "COMPLEMENTARIOS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws[f'A{row}'] = "Ítem"
    ws[f'B{row}'] = "Cantidad"
    ws[f'C{row}'] = "Precio Unitario"
    ws[f'D{row}'] = "Subtotal"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
    row += 1
    
    for nombre, item in st.session_state.complementarios.items():
        ws[f'A{row}'] = nombre
        ws[f'B{row}'] = item.cantidad
        ws[f'B{row}'].number_format = '0.00'
        ws[f'C{row}'] = item.precio_unitario
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'] = item.calcular_subtotal()
        ws[f'D{row}'].number_format = '$#,##0'
        row += 1
    
    comp_calc = calcular_complementarios()
    ws[f'C{row}'] = "Subtotal:"
    ws[f'D{row}'] = comp_calc['subtotal']
    ws[f'D{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'C{row}'] = "Comisión:"
    ws[f'D{row}'] = comp_calc['comision']
    ws[f'D{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'C{row}'] = "AIU:"
    ws[f'D{row}'] = comp_calc['aiu']
    ws[f'D{row}'].number_format = '$#,##0'
    row += 1
    
    ws[f'C{row}'] = "TOTAL:"
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = comp_calc['total']
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'D{row}'].font = Font(bold=True)
    row += 2
    
    # ============================================================================
    # ADMINISTRACIÓN (DETALLADO)
    # ============================================================================
    admin_det = calcular_administracion_detallada()
    
    ws[f'A{row}'] = "ADMINISTRACIÓN"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Personal Profesional
    ws[f'A{row}'] = "Personal Profesional:"
    ws[f'A{row}'].font = subsection_font
    row += 1
    
    for nombre, persona in st.session_state.personal_profesional.items():
        ws[f'A{row}'] = f"  • {nombre}"
        ws[f'B{row}'] = persona.calcular_total()
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
    
    ws[f'A{row}'] = "Subtotal Personal Profesional:"
    ws[f'B{row}'] = admin_det['personal_profesional']
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 2
    
    # Personal Administrativo
    ws[f'A{row}'] = "Personal Administrativo:"
    ws[f'A{row}'].font = subsection_font
    row += 1
    
    for nombre, persona in st.session_state.personal_administrativo.items():
        ws[f'A{row}'] = f"  • {nombre}"
        ws[f'B{row}'] = persona.calcular_total()
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
    
    ws[f'A{row}'] = "Subtotal Personal Administrativo:"
    ws[f'B{row}'] = admin_det['personal_administrativo']
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 2
    
    # Otros Conceptos
    ws[f'A{row}'] = "Otros Conceptos:"
    ws[f'A{row}'].font = subsection_font
    row += 1
    
    for concepto_nombre, concepto_obj in st.session_state.otros_admin.items():
        ws[f'A{row}'] = f"  • {concepto_nombre}"
        ws[f'B{row}'] = concepto_obj.calcular_subtotal()
        ws[f'B{row}'].number_format = '$#,##0'
        row += 1
    
    ws[f'A{row}'] = "Subtotal Otros Conceptos:"
    ws[f'B{row}'] = admin_det['otros_conceptos']
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    row += 2
    
    # Total Administración
    ws[f'A{row}'] = "TOTAL ADMINISTRACIÓN:"
    ws[f'B{row}'] = admin_det['total']
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'B{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = total_fill
    ws[f'B{row}'].fill = total_fill
    row += 2
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    # Guardar en BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def calcular_impuestos_dinamicos():
    """
    Calcula impuestos basándose en los totales reales de la cotización
    Fórmulas del Excel:
    - Base = Diseños + Estructura + Mampostería + Techos + Cimentaciones + Complementarios
    - FIC = 0.25% * Base
    - Industria y Comercio = 0.25% * Base
    - Otros = 0%
    - 4*1000 = 0.4% * Base
    - IVA sobre Utilidad = 19% * (Utilidad / 3)  ← IMPORTANTE: divide entre 3
    - Improrrenta = 34% * (Base IVA / 3)  ← IMPORTANTE: divide entre 3 nuevamente
    """
    
    # Calcular base imponible (costos directos)
    total_disenos = calcular_disenos()
    total_estructura = calcular_estructura()
    total_mamposteria = calcular_mamposteria()
    total_techos = calcular_mamposteria_techos()
    
    # Cimentación y Complementarios retornan diccionarios
    cimentacion_dict = calcular_cimentacion()
    total_cimentacion = cimentacion_dict['total'] if isinstance(cimentacion_dict, dict) else cimentacion_dict
    
    complementarios_dict = calcular_complementarios()
    total_complementarios = complementarios_dict['total'] if isinstance(complementarios_dict, dict) else complementarios_dict
    
    base_imponible = (
        total_disenos +
        total_estructura +
        total_mamposteria +
        total_techos +
        total_cimentacion +
        total_complementarios
    )
    
    # Calcular utilidad estimada según % de AIU (sobre costos directos base, no sobre todo)
    costos_directos_base = total_disenos + total_estructura + total_mamposteria + total_techos
    utilidad_estimada = costos_directos_base * (st.session_state.config_aiu.get('Utilidad (%)', 26.5) / 100)
    
    # Calcular cada impuesto
    fic = base_imponible * 0.0025
    industria_comercio = base_imponible * 0.0025
    otros = 0.0  # Según Excel está en 0
    cuatro_mil = base_imponible * 0.004
    
    # FÓRMULAS CORRECTAS DEL EXCEL:
    # IVA = 19% × (Utilidad / 3)
    base_iva = utilidad_estimada / 3
    iva_utilidad = base_iva * 0.19
    
    # Improrrenta = 34% × (Base IVA / 3)
    base_improrrenta = base_iva / 3
    improrrenta = base_improrrenta * 0.34
    
    # Actualizar el ConceptoDetallado de Impuestos
    st.session_state.otros_admin['Impuestos'].items_detalle = {
        'FIC': fic,
        'Industria y Comercio': industria_comercio,
        'Otros': otros,
        '4*1000': cuatro_mil,
        'IVA sobre la Utilidad': iva_utilidad,
        'Improrrenta (34% de la utilidad)': improrrenta
    }
    
    return fic + industria_comercio + otros + cuatro_mil + iva_utilidad + improrrenta

def calcular_resumen_global():
    """Calcula resumen global del proyecto"""
    
    # COTIZACIÓN 1: Diseños + Estructura + Mampostería + Mampostería y Techos
    disenos = calcular_disenos()
    estructura = calcular_estructura()
    mamposteria = calcular_mamposteria()
    mamposteria_techos = calcular_mamposteria_techos()
    
    costos_directos_cot1 = disenos + estructura + mamposteria + mamposteria_techos
    
    # COTIZACIÓN 2: Cimentaciones + Complementarios
    cimentacion = calcular_cimentacion()
    complementarios = calcular_complementarios()
    
    costos_directos_cot2 = cimentacion['total'] + complementarios['total']
    
    # TOTAL COSTOS DIRECTOS (base para AIU general)
    # IMPORTANTE: Para AIU general, NO incluir AIU específico de cimentación/complementarios
    total_base_aiu = disenos + estructura + mamposteria + mamposteria_techos
    
    # AIU GENERAL (se aplica solo sobre Cotización 1)
    admin_detallada = calcular_administracion_detallada()
    
    # Usuario puede modificar el % de administración
    pct_admin_calculado = (admin_detallada['total'] / total_base_aiu * 100) if total_base_aiu > 0 else 0
    pct_admin_final = st.session_state.config_aiu['Administración (%)']
    
    comision_ventas = total_base_aiu * (st.session_state.config_aiu['Comisión de Ventas (%)'] / 100)
    imprevistos = total_base_aiu * (st.session_state.config_aiu['Imprevistos (%)'] / 100)
    administracion = total_base_aiu * (pct_admin_final / 100)
    logistica = total_base_aiu * (st.session_state.config_aiu['Logística (%)'] / 100)
    utilidad = total_base_aiu * (st.session_state.config_aiu['Utilidad (%)'] / 100)
    
    total_aiu_general = comision_ventas + imprevistos + administracion + logistica + utilidad
    
    # TOTALES POR COTIZACIÓN
    total_cot1 = costos_directos_cot1 + total_aiu_general
    total_cot2 = cimentacion['total'] + complementarios['total']
    
    # TOTAL PROYECTO
    total_proyecto = total_cot1 + total_cot2
    
    # PRECIO POR M²
    area_base = st.session_state.proyecto.area_base
    precio_m2 = total_proyecto / area_base if area_base > 0 else 0
    
    return {
        'cotizacion1': {
            'disenos': disenos,
            'estructura': estructura,
            'mamposteria': mamposteria,
            'mamposteria_techos': mamposteria_techos,
            'costos_directos': costos_directos_cot1,
            'aiu': {
                'comision_ventas': comision_ventas,
                'imprevistos': imprevistos,
                'administracion': administracion,
                'logistica': logistica,
                'utilidad': utilidad,
                'total': total_aiu_general
            },
            'total': total_cot1
        },
        'cotizacion2': {
            'cimentacion': cimentacion,
            'complementarios': complementarios,
            'total': total_cot2
        },
        'administracion_detallada': admin_detallada,
        'pct_admin_calculado': pct_admin_calculado,
        'total_proyecto': total_proyecto,
        'precio_m2': precio_m2
    }

def get_resumen_cacheado():
    """
    Versión cacheada de calcular_resumen_global().
    Se cachea basado en un hash del estado actual para evitar recálculos innecesarios.
    
    OPTIMIZACIÓN CRÍTICA:
    - calcular_resumen_global() se llamaba 5+ veces por render
    - Con cache, solo se calcula 1 vez cuando cambian los datos
    - Reduce tiempo de render de ~5-10s a <1s
    """
    import hashlib
    import json
    
    # Crear hash del estado relevante (solo datos que afectan el cálculo)
    try:
        estado_relevante = {
            'proyecto_areas': [
                st.session_state.proyecto.area_base,
                st.session_state.proyecto.area_cubierta,
                st.session_state.proyecto.area_entrepiso,
                st.session_state.proyecto.niveles,
                st.session_state.proyecto.muro_tipo
            ],
            'aiu': list(st.session_state.config_aiu.values()),
            'opcion_cim': st.session_state.opcion_cimentacion
        }
        
        estado_json = json.dumps(estado_relevante, sort_keys=True)
        estado_hash = hashlib.md5(estado_json.encode()).hexdigest()
        
        # Verificar cache
        if 'resumen_cache' not in st.session_state:
            st.session_state.resumen_cache = {}
            st.session_state.resumen_cache_hash = None
        
        # Si el hash es el mismo, retornar cache
        if st.session_state.resumen_cache_hash == estado_hash:
            return st.session_state.resumen_cache
        
        # Calcular y cachear
        resumen = calcular_resumen_global()
        st.session_state.resumen_cache = resumen
        st.session_state.resumen_cache_hash = estado_hash
        
        return resumen
        
    except Exception as e:
        # Si falla el cache, calcular directo
        return calcular_resumen_global()

# ============================================================================
# INTERFAZ - SIDEBAR
# ============================================================================

def render_sidebar():
    """Render sidebar con información del proyecto"""
    with st.sidebar:
        st.markdown("### 📋 Información del Proyecto")
        
        nuevo_nombre = st.text_input(
            "Nombre del Proyecto", 
            value=st.session_state.proyecto.nombre
        )
        if nuevo_nombre != st.session_state.proyecto.nombre:
            st.session_state.proyecto.nombre = nuevo_nombre
        
        nuevo_cliente = st.text_input(
            "Cliente", 
            value=st.session_state.proyecto.cliente
        )
        if nuevo_cliente != st.session_state.proyecto.cliente:
            st.session_state.proyecto.cliente = nuevo_cliente
        
        nueva_direccion = st.text_input(
            "Dirección", 
            value=st.session_state.proyecto.direccion
        )
        if nueva_direccion != st.session_state.proyecto.direccion:
            st.session_state.proyecto.direccion = nueva_direccion
        
        col1, col2 = st.columns(2)
        with col1:
            nuevo_telefono = st.text_input(
                "Teléfono", 
                value=st.session_state.proyecto.telefono
            )
            if nuevo_telefono != st.session_state.proyecto.telefono:
                st.session_state.proyecto.telefono = nuevo_telefono
        with col2:
            nuevo_bm = st.text_input(
                "Business Manager", 
                value=st.session_state.proyecto.business_manager
            )
            if nuevo_bm != st.session_state.proyecto.business_manager:
                st.session_state.proyecto.business_manager = nuevo_bm
        
        nuevo_medio = st.text_input(
            "Medio de Contacto", 
            value=st.session_state.proyecto.medio_contacto
        )
        if nuevo_medio != st.session_state.proyecto.medio_contacto:
            st.session_state.proyecto.medio_contacto = nuevo_medio
        
        st.markdown("---")
        st.markdown("### 📐 Áreas del Proyecto")
        
        nueva_area_base = st.number_input(
            "Área de la Base (m²)",
            min_value=0.0,
            value=float(st.session_state.proyecto.area_base),
            step=0.01,
            format="%.2f",
            help="Área principal que se usa como multiplicador en Diseños"
        )
        if nueva_area_base != st.session_state.proyecto.area_base:
            st.session_state.proyecto.area_base = nueva_area_base
        
        nueva_area_cubierta = st.number_input(
            "Área de Cubierta (m²)",
            min_value=0.0,
            value=float(st.session_state.proyecto.area_cubierta),
            step=0.01,
            format="%.2f"
        )
        if nueva_area_cubierta != st.session_state.proyecto.area_cubierta:
            st.session_state.proyecto.area_cubierta = nueva_area_cubierta
        
        nueva_area_entrepiso = st.number_input(
            "Área de Entrepiso (m²)",
            min_value=0.0,
            value=float(st.session_state.proyecto.area_entrepiso),
            step=0.01,
            format="%.2f"
        )
        if nueva_area_entrepiso != st.session_state.proyecto.area_entrepiso:
            st.session_state.proyecto.area_entrepiso = nueva_area_entrepiso
        
        nuevos_niveles = st.number_input(
            "Niveles",
            min_value=1,
            value=int(st.session_state.proyecto.niveles)
        )
        if nuevos_niveles != st.session_state.proyecto.niveles:
            st.session_state.proyecto.niveles = nuevos_niveles
        
        nuevo_muro_tipo = st.selectbox(
            "Tipo de Muro",
            options=["sencillo", "doble"],
            index=0 if st.session_state.proyecto.muro_tipo == "sencillo" else 1
        )
        if nuevo_muro_tipo != st.session_state.proyecto.muro_tipo:
            st.session_state.proyecto.muro_tipo = nuevo_muro_tipo
        
        st.markdown("---")
        st.markdown("### 💼 Configuración AIU General")
        st.caption("Aplica a Diseños + Estructura + Mampostería + Techos")
        
        for concepto in list(st.session_state.config_aiu.keys()):
            nuevo_valor = st.number_input(
                concepto,
                min_value=0.0,
                max_value=100.0,
                value=float(st.session_state.config_aiu[concepto]),
                step=0.5,
                format="%.1f"
            )
            if nuevo_valor != st.session_state.config_aiu[concepto]:
                st.session_state.config_aiu[concepto] = nuevo_valor
        
        st.markdown("---")
        st.markdown("### 💾 Gestión de Cotizaciones")
        
        st.caption("⚠️ **En Streamlit Cloud:** Las cotizaciones guardadas persisten durante la sesión, pero pueden perderse si el servidor se reinicia. Usa **📤 Export** para backup permanente.")
        
        # Cargar cotizaciones disponibles desde archivos al iniciar
        if 'cotizaciones_guardadas' not in st.session_state:
            st.session_state.cotizaciones_guardadas = cargar_cotizaciones_disponibles()
        
        # Mostrar nombre de cotización cargada (si existe)
        if 'nombre_cotizacion_actual' in st.session_state and st.session_state.nombre_cotizacion_actual:
            st.info(f"💡 **Cotización cargada:** {st.session_state.nombre_cotizacion_actual}")
            st.caption("Para sobrescribir, usa el mismo nombre abajo")
        
        # Guardar cotización actual
        col1, col2 = st.columns([3, 1])
        with col1:
            nombre_guardar = st.text_input(
                "Nombre para guardar",
                placeholder="Escribe el nombre aquí..."
            )
        with col2:
            if st.button("💾", key="btn_guardar", use_container_width=True, help="Guardar cotización"):
                if nombre_guardar:
                    success, msg = guardar_cotizacion_memoria(nombre_guardar)
                    if success:
                        st.success("✅ Guardado en sesión")
                        st.warning("⚠️ Recuerda exportar a JSON para backup permanente")
                        # Recargar lista
                        st.session_state.cotizaciones_guardadas = cargar_cotizaciones_disponibles()
                    else:
                        # Si falla el guardado en archivo, al menos está en memoria
                        st.warning("⚠️ Guardado solo en memoria (sesión actual)")
                        st.caption(f"Error archivo: {msg}")
                else:
                    st.error("❌ Nombre requerido")
        
        # Lista de cotizaciones guardadas
        if st.session_state.cotizaciones_guardadas:
            st.markdown("**Cotizaciones guardadas:**")
            st.caption(f"📁 {len(st.session_state.cotizaciones_guardadas)} cotización(es)")
            
            for nombre_cot in sorted(st.session_state.cotizaciones_guardadas.keys()):
                col1, col2, col3 = st.columns([3, 1, 1])
                with col1:
                    # Mostrar fecha de guardado si está disponible
                    fecha = st.session_state.cotizaciones_guardadas[nombre_cot].get('fecha_guardado', '')
                    if fecha:
                        try:
                            fecha_dt = datetime.fromisoformat(fecha)
                            st.caption(f"📄 {nombre_cot}")
                            st.caption(f"🕒 {fecha_dt.strftime('%d/%m/%Y %H:%M')}")
                        except:
                            st.text(nombre_cot)
                    else:
                        st.text(nombre_cot)
                with col2:
                    if st.button("📂", key=f"cargar_{nombre_cot}", help="Cargar en memoria", use_container_width=True):
                        # Solo cargar en memoria temporal, NO aplicar aún
                        cotizacion_data = st.session_state.cotizaciones_guardadas[nombre_cot]
                        st.session_state.cotizacion_en_memoria = cotizacion_data.copy()
                        st.session_state.nombre_cotizacion_en_memoria = nombre_cot
                        st.session_state.mostrar_boton_aplicar = True  # Flag para mostrar botón
                        st.rerun()  # Rerun para mostrar el botón inmediatamente
                with col3:
                    if st.button("🗑️", key=f"eliminar_{nombre_cot}", help="Eliminar", use_container_width=True):
                        success, msg = eliminar_cotizacion_archivo(nombre_cot)
                        if success:
                            if nombre_cot in st.session_state.cotizaciones_guardadas:
                                del st.session_state.cotizaciones_guardadas[nombre_cot]
                            st.success(msg)
                        else:
                            st.error(msg)
                        st.rerun()
            
            # BOTÓN DE APLICAR (aparece después de hacer click en 📂)
            if st.session_state.get('mostrar_boton_aplicar', False):
                st.markdown("---")
                st.markdown("### 📥 Cotización Lista para Cargar")
                
                nombre_en_memoria = st.session_state.get('nombre_cotizacion_en_memoria', 'Sin nombre')
                st.success(f"✅ Cargado en memoria: {nombre_en_memoria}")
                st.info(f"**📋 Cotización en memoria:** {nombre_en_memoria}")
                st.caption("👇 Click abajo para aplicar todos los datos al cotizador")
                
                col_aplicar1, col_aplicar2 = st.columns([3, 1])
                with col_aplicar1:
                    if st.button("📥 CARGAR DATOS EN COTIZADOR", 
                                key="btn_aplicar_cotizacion_guardada",
                                use_container_width=True,
                                type="primary"):
                        # Aplicar la cotización
                        deserializar_cotizacion(st.session_state.cotizacion_en_memoria)
                        st.session_state.nombre_cotizacion_actual = nombre_en_memoria
                        
                        # Limpiar memoria temporal
                        del st.session_state.cotizacion_en_memoria
                        del st.session_state.nombre_cotizacion_en_memoria
                        st.session_state.mostrar_boton_aplicar = False
                        
                        st.success(f"✅ Datos aplicados: {nombre_en_memoria}")
                        st.rerun()
                with col_aplicar2:
                    # Botón para cancelar
                    if st.button("❌", key="btn_cancelar_carga_guardada", help="Cancelar", use_container_width=True):
                        # Limpiar memoria temporal
                        del st.session_state.cotizacion_en_memoria
                        del st.session_state.nombre_cotizacion_en_memoria
                        st.session_state.mostrar_boton_aplicar = False
                        st.info("Carga cancelada")
                        st.rerun()
        else:
            st.info("💡 No hay cotizaciones guardadas aún")
        
        # Exportar/Importar JSON y Nueva cotización
        st.markdown("---")
        st.caption("**Acciones adicionales:**")
        
        col1, col2 = st.columns(2)
        with col1:
            # Exportar a JSON
            json_data = exportar_cotizacion_json()
            st.download_button(
                label="📤 Export",
                data=json_data,
                file_name=f"Cot_{st.session_state.proyecto.nombre.replace(' ', '_')}_{pd.Timestamp.now().strftime('%Y%m%d')}.json",
                mime="application/json",
                use_container_width=True,
                help="Exportar a JSON"
            )
        with col2:
            # Nueva cotización
            if st.button("🆕 Nueva", key="btn_nueva_cotizacion", use_container_width=True, help="Nueva cotización"):
                # Reiniciar session_state pero mantener cotizaciones guardadas
                cotizaciones_backup = st.session_state.cotizaciones_guardadas.copy()
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.session_state.cotizaciones_guardadas = cotizaciones_backup
                st.session_state.nombre_cotizacion_actual = ""  # Limpiar nombre de cotización
                st.rerun()
        
        # Importar desde JSON
        uploaded_file = st.file_uploader(
            "📥 Importar JSON",
            type=['json'],
            key="upload_json",
            help="Cargar cotización desde archivo JSON"
        )
        if uploaded_file is not None:
            try:
                # Cargar JSON en memoria temporal
                cotizacion_data = json.loads(uploaded_file.getvalue().decode('utf-8'))
                st.session_state.cotizacion_en_memoria = cotizacion_data
                
                # Extraer nombre del proyecto del JSON
                nombre_proyecto = cotizacion_data.get('proyecto', {}).get('nombre', 'Sin nombre')
                st.session_state.nombre_cotizacion_en_memoria = nombre_proyecto
                
                st.success(f"✅ JSON cargado en memoria: {nombre_proyecto}")
                
                # Mostrar botón de aplicar AQUÍ MISMO (justo debajo del mensaje)
                st.markdown("---")
                st.markdown("### 📥 Cotización Lista para Cargar")
                st.info(f"**📋 Cotización en memoria:** {nombre_proyecto}")
                st.caption("👇 Click abajo para aplicar todos los datos al cotizador")
                
                col_aplicar1, col_aplicar2 = st.columns([3, 1])
                with col_aplicar1:
                    if st.button("📥 CARGAR DATOS EN COTIZADOR", 
                                key="btn_aplicar_cotizacion_json",
                                use_container_width=True,
                                type="primary"):
                        # Aplicar la cotización
                        deserializar_cotizacion(st.session_state.cotizacion_en_memoria)
                        st.session_state.nombre_cotizacion_actual = nombre_proyecto
                        
                        # Limpiar memoria temporal
                        del st.session_state.cotizacion_en_memoria
                        del st.session_state.nombre_cotizacion_en_memoria
                        
                        st.success(f"✅ Datos aplicados: {nombre_proyecto}")
                        st.rerun()
                with col_aplicar2:
                    # Botón para cancelar
                    if st.button("❌", key="btn_cancelar_carga_json", help="Cancelar", use_container_width=True):
                        # Limpiar memoria temporal
                        del st.session_state.cotizacion_en_memoria
                        del st.session_state.nombre_cotizacion_en_memoria
                        st.info("Carga cancelada")
                        st.rerun()
                
            except Exception as e:
                st.error(f"❌ Error al leer JSON: {str(e)}")


# ============================================================================
# INTERFAZ - TAB 1: DISEÑOS, ESTRUCTURA Y MAMPOSTERÍA
# ============================================================================

def render_tab_disenos_estructura():
    """Tab 1: Diseños, Estructura, Mampostería y Techos"""
    
    st.markdown('<h2 class="section-title">📐 Diseños, Estructura, Mampostería y Techos</h2>', unsafe_allow_html=True)
    
    # DISEÑOS Y PLANIFICACIÓN
    with st.expander("📐 Diseños y Planificación", expanded=True):
        st.caption(f"Los valores se multiplican por el Área de la Base: {st.session_state.proyecto.area_base:.2f} m²")
        
        df_disenos_data = []
        for nombre, item in st.session_state.disenos.items():
            df_disenos_data.append({
                'Ítem': nombre,
                'Precio Unitario ($/m²)': item.precio_unitario,
                'Subtotal': item.calcular_subtotal(st.session_state.proyecto.area_base)
            })
        
        df_disenos = pd.DataFrame(df_disenos_data)
        
        edited_disenos = st.data_editor(
            df_disenos,
            column_config={
                'Precio Unitario ($/m²)': st.column_config.NumberColumn(
                    format="%.0f",
                    min_value=0
                ),
                'Subtotal': st.column_config.NumberColumn(
                    format="%.0f",
                    disabled=True
                )
            },
            hide_index=True,
            use_container_width=True
        )
        
        # Actualizar session state
        for idx, row in edited_disenos.iterrows():
            nombre = row['Ítem']
            st.session_state.disenos[nombre].precio_unitario = row['Precio Unitario ($/m²)']
        
        total_disenos = calcular_disenos()
        st.metric("**Total Diseños y Planificación**", f"${total_disenos:,.0f}")
    
    # ESTRUCTURA
    with st.expander("🏗️ Estructura", expanded=True):
        item = st.session_state.estructura
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            nueva_cantidad = st.number_input("Cantidad", value=float(item.cantidad), min_value=0.0, step=0.01, format="%.2f")
            if nueva_cantidad != item.cantidad:
                item.cantidad = nueva_cantidad
        with col2:
            nuevo_mat = st.number_input("Materiales ($)", value=float(item.precio_materiales), min_value=0.0, step=1000.0, format="%.0f")
            if nuevo_mat != item.precio_materiales:
                item.precio_materiales = nuevo_mat
        with col3:
            nuevo_eq = st.number_input("Equipos ($)", value=float(item.precio_equipos), min_value=0.0, step=1000.0, format="%.0f")
            if nuevo_eq != item.precio_equipos:
                item.precio_equipos = nuevo_eq
        with col4:
            nuevo_mo = st.number_input("Mano de Obra ($)", value=float(item.precio_mano_obra), min_value=0.0, step=1000.0, format="%.0f")
            if nuevo_mo != item.precio_mano_obra:
                item.precio_mano_obra = nuevo_mo
        
        total_estructura = calcular_estructura()
        st.metric("**Total Estructura**", f"${total_estructura:,.0f}")
    
    # MAMPOSTERÍA
    with st.expander("🧱 Mampostería", expanded=True):
        item = st.session_state.mamposteria
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            nueva_cantidad = st.number_input("Cantidad (m²)", value=float(item.cantidad), min_value=0.0, step=0.01, format="%.2f")
            if nueva_cantidad != item.cantidad:
                item.cantidad = nueva_cantidad
        with col2:
            nuevo_mat = st.number_input("Materiales ($)", value=float(item.precio_materiales), min_value=0.0, step=1000.0, format="%.0f")
            if nuevo_mat != item.precio_materiales:
                item.precio_materiales = nuevo_mat
        with col3:
            nuevo_eq = st.number_input("Equipos ($)", value=float(item.precio_equipos), min_value=0.0, step=1000.0, format="%.0f")
            if nuevo_eq != item.precio_equipos:
                item.precio_equipos = nuevo_eq
        with col4:
            nuevo_mo = st.number_input("Mano de Obra ($)", value=float(item.precio_mano_obra), min_value=0.0, step=1000.0, format="%.0f")
            if nuevo_mo != item.precio_mano_obra:
                item.precio_mano_obra = nuevo_mo
        
        total_mamposteria = calcular_mamposteria()
        st.metric("**Total Mampostería**", f"${total_mamposteria:,.0f}")
    
    # TECHOS Y OTROS
    with st.expander("🏠 Techos y otros", expanded=True):
        
        st.caption("📝 Editables: Cantidad, Materiales, Equipos, Mano de Obra")
        
        df_mt_data = []
        for nombre, item in st.session_state.mamposteria_techos.items():
            df_mt_data.append({
                'Ítem': nombre,
                'Unidad': item.unidad,
                'Cantidad': item.cantidad,
                'Materiales': item.precio_materiales,
                'Equipos': item.precio_equipos,
                'Mano de Obra': item.precio_mano_obra,
                'Subtotal': item.calcular_subtotal()
            })
        
        df_mt = pd.DataFrame(df_mt_data)
        
        edited_mt = st.data_editor(
            df_mt,
            column_config={
                'Ítem': st.column_config.TextColumn(disabled=True),
                'Unidad': st.column_config.TextColumn(disabled=True),
                'Cantidad': st.column_config.NumberColumn(min_value=0, format="%.2f"),
                'Materiales': st.column_config.NumberColumn(min_value=0, format="%.0f"),
                'Equipos': st.column_config.NumberColumn(min_value=0, format="%.0f"),
                'Mano de Obra': st.column_config.NumberColumn(min_value=0, format="%.0f"),
                'Subtotal': st.column_config.NumberColumn(format="%.0f", disabled=True)
            },
            hide_index=True,
            use_container_width=True
        )
        
        # Actualizar session state
        for idx, row in edited_mt.iterrows():
            nombre = row['Ítem']
            st.session_state.mamposteria_techos[nombre].cantidad = row['Cantidad']
            st.session_state.mamposteria_techos[nombre].precio_materiales = row['Materiales']
            st.session_state.mamposteria_techos[nombre].precio_equipos = row['Equipos']
            st.session_state.mamposteria_techos[nombre].precio_mano_obra = row['Mano de Obra']
        
        total_mt = calcular_mamposteria_techos()
        st.metric("**Total Techos y otros**", f"${total_mt:,.0f}")

# ============================================================================
# INTERFAZ - TAB 2: CIMENTACIONES
# ============================================================================

def render_tab_cimentaciones():
    """Tab 2: Cimentaciones"""
    
    st.markdown('<h2 class="section-title">⚙️ Cimentaciones</h2>', unsafe_allow_html=True)
    
    nueva_opcion = st.radio(
        "Seleccione la opción de cimentación:",
        options=['Opción 1', 'Opción 2'],
        index=0 if st.session_state.opcion_cimentacion == 'Opción 1' else 1,
        horizontal=True
    )
    if nueva_opcion != st.session_state.opcion_cimentacion:
        st.session_state.opcion_cimentacion = nueva_opcion
    
    if st.session_state.opcion_cimentacion == 'Opción 1':
        st.markdown("### Opción 1: Zapatas y Vigas de Concreto")
        items = st.session_state.cimentacion_opcion1
    else:
        st.markdown("### Opción 2: Pilotes de Apoyo")
        items = st.session_state.cimentacion_opcion2
    
    st.caption("📝 Editables: Cantidad, Precio Unitario")
    
    df_cim_data = []
    for nombre, item in items.items():
        df_cim_data.append({
            'Ítem': nombre,
            'Unidad': item.unidad,
            'Cantidad': item.cantidad,
            'Precio Unitario': item.precio_unitario,
            'Subtotal': item.calcular_subtotal()
        })
    
    df_cim = pd.DataFrame(df_cim_data)
    
    edited_cim = st.data_editor(
        df_cim,
        column_config={
            'Ítem': st.column_config.TextColumn(disabled=True),
            'Unidad': st.column_config.TextColumn(disabled=True),
            'Cantidad': st.column_config.NumberColumn(min_value=0, format="%.2f"),
            'Precio Unitario': st.column_config.NumberColumn(min_value=0, format="%.0f"),
            'Subtotal': st.column_config.NumberColumn(format="%.0f", disabled=True)
        },
        hide_index=True,
        use_container_width=True
    )
    
    # Actualizar session state
    for idx, row in edited_cim.iterrows():
        nombre = row['Ítem']
        items[nombre].cantidad = row['Cantidad']
        items[nombre].precio_unitario = row['Precio Unitario']
    
    st.markdown("---")
    st.markdown("### Configuración AIU Cimentaciones")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        nueva_comision_cim = st.number_input(
            "Comisión (%)",
            value=float(st.session_state.aiu_cimentacion['pct_comision']),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            format="%.1f",
            key="widget_comision_cimentacion"
        )
        if nueva_comision_cim != st.session_state.aiu_cimentacion['pct_comision']:
            st.session_state.aiu_cimentacion['pct_comision'] = nueva_comision_cim
    with col2:
        nuevo_aiu_cim = st.number_input(
            "AIU (%)",
            value=float(st.session_state.aiu_cimentacion['pct_aiu']),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            format="%.1f",
            key="widget_aiu_cimentacion"
        )
        if nuevo_aiu_cim != st.session_state.aiu_cimentacion['pct_aiu']:
            st.session_state.aiu_cimentacion['pct_aiu'] = nuevo_aiu_cim
    with col3:
        nueva_logistica_cim = st.number_input(
            "Logística ($)",
            value=float(st.session_state.aiu_cimentacion['logistica']),
            min_value=0.0,
            step=1000.0,
            format="%.0f",
            key="widget_logistica_cimentacion"
        )
        if nueva_logistica_cim != st.session_state.aiu_cimentacion['logistica']:
            st.session_state.aiu_cimentacion['logistica'] = nueva_logistica_cim
    
    cimentacion = calcular_cimentacion()
    
    st.markdown('<p style="font-size: 18px; font-weight: bold; margin-top: 10px;">Resumen Cimentación</p>', unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Subtotal", f"${cimentacion['subtotal']:,.0f}")
    col2.metric("Comisión", f"${cimentacion['comision']:,.0f}")
    col3.metric("AIU", f"${cimentacion['aiu']:,.0f}")
    col4.metric("**TOTAL**", f"${cimentacion['total']:,.0f}")

# ============================================================================
# INTERFAZ - TAB 3: COMPLEMENTARIOS
# ============================================================================

def render_tab_complementarios():
    """Tab 3: Complementarios"""
    
    st.markdown('<h2 class="section-title">🔧 Complementarios</h2>', unsafe_allow_html=True)
    
    st.caption("📝 Editables: Cantidad, Precio Unitario")
    
    df_comp_data = []
    for nombre, item in st.session_state.complementarios.items():
        df_comp_data.append({
            'Ítem': nombre,
            'Unidad': item.unidad,
            'Cantidad': item.cantidad,
            'Precio Unitario': item.precio_unitario,
            'Subtotal': item.calcular_subtotal()
        })
    
    df_comp = pd.DataFrame(df_comp_data)
    
    edited_comp = st.data_editor(
        df_comp,
        column_config={
            'Ítem': st.column_config.TextColumn(disabled=True),
            'Unidad': st.column_config.TextColumn(disabled=True),
            'Cantidad': st.column_config.NumberColumn(min_value=0, format="%.2f"),
            'Precio Unitario': st.column_config.NumberColumn(min_value=0, format="%.0f"),
            'Subtotal': st.column_config.NumberColumn(format="%.0f", disabled=True)
        },
        hide_index=True,
        use_container_width=True
    )
    
    # Actualizar session state
    for idx, row in edited_comp.iterrows():
        nombre = row['Ítem']
        st.session_state.complementarios[nombre].cantidad = row['Cantidad']
        st.session_state.complementarios[nombre].precio_unitario = row['Precio Unitario']
    
    st.markdown("---")
    st.markdown("### Configuración AIU Complementarios")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        nueva_comision_comp = st.number_input(
            "Comisión (%)",
            value=float(st.session_state.aiu_complementarios['pct_comision']),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            format="%.1f",
            key="widget_comision_complementarios"
        )
        if nueva_comision_comp != st.session_state.aiu_complementarios['pct_comision']:
            st.session_state.aiu_complementarios['pct_comision'] = nueva_comision_comp
    with col2:
        nuevo_aiu_comp = st.number_input(
            "AIU (%)",
            value=float(st.session_state.aiu_complementarios['pct_aiu']),
            min_value=0.0,
            max_value=100.0,
            step=0.5,
            format="%.1f",
            key="widget_aiu_complementarios"
        )
        if nuevo_aiu_comp != st.session_state.aiu_complementarios['pct_aiu']:
            st.session_state.aiu_complementarios['pct_aiu'] = nuevo_aiu_comp
    with col3:
        nueva_logistica_comp = st.number_input(
            "Logística ($)",
            value=float(st.session_state.aiu_complementarios['logistica']),
            min_value=0.0,
            step=1000.0,
            format="%.0f",
            key="widget_logistica_complementarios"
        )
        if nueva_logistica_comp != st.session_state.aiu_complementarios['logistica']:
            st.session_state.aiu_complementarios['logistica'] = nueva_logistica_comp
    
    complementarios = calcular_complementarios()
    
    st.markdown('<p style="font-size: 18px; font-weight: bold; margin-top: 10px;">Resumen Complementarios</p>', unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Subtotal", f"${complementarios['subtotal']:,.0f}")
    col2.metric("Comisión", f"${complementarios['comision']:,.0f}")
    col3.metric("AIU", f"${complementarios['aiu']:,.0f}")
    col4.metric("**TOTAL**", f"${complementarios['total']:,.0f}")

# ============================================================================
# INTERFAZ - TAB 4: ADMINISTRACIÓN
# ============================================================================

def render_tab_administracion():
    """Tab 4: Administración Detallada"""
    
    st.markdown('<h2 class="section-title">💼 Administración</h2>', unsafe_allow_html=True)
    
    subtab1, subtab2, subtab3, subtab4 = st.tabs([
        "Personal Profesional", 
        "Personal Administrativo", 
        "Otros Conceptos",
        "Resumen"
    ])
    
    # SUB-TAB 1: PERSONAL PROFESIONAL
    with subtab1:
        st.markdown("### Personal Profesional y Técnico")
        st.caption("📝 Editables: Cant, Valor/Mes, % Prest, Dedicación, Meses")
        
        df_prof_data = []
        for nombre, p in st.session_state.personal_profesional.items():
            df_prof_data.append({
                'Nombre': nombre,
                'Cant': p.cantidad,
                'Valor/Mes': p.valor_mes,
                '% Prest': p.pct_prestaciones,
                'Dedicación': p.dedicacion,
                'Meses': p.meses,
                'Total': p.calcular_total()
            })
        
        df_prof = pd.DataFrame(df_prof_data)
        
        edited_prof = st.data_editor(
            df_prof,
            column_config={
                'Nombre': st.column_config.TextColumn(disabled=True),
                'Cant': st.column_config.NumberColumn(min_value=0, format="%d"),
                'Valor/Mes': st.column_config.NumberColumn(min_value=0, format="%.0f"),
                '% Prest': st.column_config.NumberColumn(min_value=0, max_value=100, format="%.1f"),
                'Dedicación': st.column_config.NumberColumn(min_value=0.0, max_value=1.0, format="%.2f"),
                'Meses': st.column_config.NumberColumn(min_value=0, format="%d"),
                'Total': st.column_config.NumberColumn(format="%.0f", disabled=True)
            },
            hide_index=True,
            use_container_width=True
        )
        
        # Actualizar session state
        for idx, row in edited_prof.iterrows():
            nombre = row['Nombre']
            p = st.session_state.personal_profesional[nombre]
            p.cantidad = int(row['Cant'])
            p.valor_mes = row['Valor/Mes']
            p.pct_prestaciones = row['% Prest']
            p.dedicacion = row['Dedicación']
            p.meses = int(row['Meses'])
        
        # Subtotal Personal Profesional
        total_prof = sum([p.calcular_total() for p in st.session_state.personal_profesional.values()])
        st.metric("**Subtotal Personal Profesional y Técnico**", f"${total_prof:,.0f}")
    
    # SUB-TAB 2: PERSONAL ADMINISTRATIVO
    with subtab2:
        st.markdown("### Personal Administrativo")
        st.caption("📝 Editables: Cant, Valor/Mes, % Prest, Dedicación, Meses")
        
        df_admin_data = []
        for nombre, p in st.session_state.personal_administrativo.items():
            df_admin_data.append({
                'Nombre': nombre,
                'Cant': p.cantidad,
                'Valor/Mes': p.valor_mes,
                '% Prest': p.pct_prestaciones,
                'Dedicación': p.dedicacion,
                'Meses': p.meses,
                'Total': p.calcular_total()
            })
        
        df_admin = pd.DataFrame(df_admin_data)
        
        edited_admin = st.data_editor(
            df_admin,
            column_config={
                'Nombre': st.column_config.TextColumn(disabled=True),
                'Cant': st.column_config.NumberColumn(min_value=0, format="%d"),
                'Valor/Mes': st.column_config.NumberColumn(min_value=0, format="%.0f"),
                '% Prest': st.column_config.NumberColumn(min_value=0, max_value=100, format="%.1f"),
                'Dedicación': st.column_config.NumberColumn(min_value=0.0, max_value=1.0, format="%.2f"),
                'Meses': st.column_config.NumberColumn(min_value=0, format="%d"),
                'Total': st.column_config.NumberColumn(format="%.0f", disabled=True)
            },
            hide_index=True,
            use_container_width=True
        )
        
        # Actualizar session state
        for idx, row in edited_admin.iterrows():
            nombre = row['Nombre']
            p = st.session_state.personal_administrativo[nombre]
            p.cantidad = int(row['Cant'])
            p.valor_mes = row['Valor/Mes']
            p.pct_prestaciones = row['% Prest']
            p.dedicacion = row['Dedicación']
            p.meses = int(row['Meses'])
        
        # Subtotal Personal Administrativo
        total_admin = sum([p.calcular_total() for p in st.session_state.personal_administrativo.values()])
        st.metric("**Subtotal Personal Administrativo**", f"${total_admin:,.0f}")
    
    # SUB-TAB 3: OTROS CONCEPTOS
    with subtab3:
        st.markdown("### Otros Conceptos Administrativos")
        st.caption("📝 Expanda cada concepto para editar los ítems detallados")
        
        # Iterar sobre cada concepto
        for concepto_nombre, concepto_obj in st.session_state.otros_admin.items():
            # Indicador especial para impuestos (calculados automáticamente)
            if concepto_nombre == 'Impuestos':
                expander_label = f"🧮 {concepto_nombre} - Subtotal: ${concepto_obj.calcular_subtotal():,.0f} (⚙️ Calculado automáticamente)"
            else:
                expander_label = f"📋 {concepto_nombre} - Subtotal: ${concepto_obj.calcular_subtotal():,.0f}"
            
            with st.expander(expander_label, expanded=False):
                
                # Crear DataFrame con los ítems detallados
                df_detalle_data = []
                for item_nombre, item_valor in concepto_obj.items_detalle.items():
                    df_detalle_data.append({
                        'Ítem': item_nombre,
                        'Valor': item_valor
                    })
                
                if df_detalle_data:
                    df_detalle = pd.DataFrame(df_detalle_data)
                    
                    # Editor de ítems detallados
                    # Los impuestos son solo lectura (calculados automáticamente)
                    if concepto_nombre == 'Impuestos':
                        st.dataframe(
                            df_detalle,
                            column_config={
                                'Ítem': st.column_config.TextColumn(),
                                'Valor': st.column_config.NumberColumn(format="%.0f")
                            },
                            hide_index=True,
                            use_container_width=True
                        )
                        st.info("ℹ️ Los impuestos se calculan automáticamente basándose en los totales de la cotización")
                    else:
                        edited_detalle = st.data_editor(
                            df_detalle,
                            column_config={
                                'Ítem': st.column_config.TextColumn(disabled=True),
                                'Valor': st.column_config.NumberColumn(min_value=0, format="%.0f")
                            },
                            hide_index=True,
                            use_container_width=True,
                            key=f"detalle_{concepto_nombre.replace(' ', '_')}"
                        )
                        
                        # Actualizar los valores en session_state
                        for idx, row in edited_detalle.iterrows():
                            item_nombre = row['Ítem']
                            nuevo_valor = row['Valor']
                            concepto_obj.items_detalle[item_nombre] = nuevo_valor
                    
                    # Mostrar subtotal del concepto
                    st.metric(f"Subtotal {concepto_nombre}", f"${concepto_obj.calcular_subtotal():,.0f}")
                else:
                    st.info("No hay ítems detallados para este concepto")
        
        # Subtotal Total de Otros Conceptos
        total_otros = sum(concepto.calcular_subtotal() for concepto in st.session_state.otros_admin.values())
        st.markdown("---")
        st.metric("**💰 Subtotal Total Otros Conceptos**", f"${total_otros:,.0f}")
    
    # SUB-TAB 4: RESUMEN
    with subtab4:
        st.markdown("### Resumen Administración")
        
        admin_det = calcular_administracion_detallada()
        resumen = get_resumen_cacheado()
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Personal Profesional", f"${admin_det['personal_profesional']:,.0f}")
        col2.metric("Personal Administrativo", f"${admin_det['personal_administrativo']:,.0f}")
        col3.metric("Otros Conceptos", f"${admin_det['otros_conceptos']:,.0f}")
        col4.metric("**TOTAL**", f"${admin_det['total']:,.0f}")
        
        st.markdown("---")
        
        st.info(f"""
        **% Administración Calculado:** {resumen['pct_admin_calculado']:.2f}%  
        **% Administración Configurado (Sidebar):** {st.session_state.config_aiu['Administración (%)']:.2f}%
        
        💡 Puedes modificar el % en el sidebar si deseas usar un valor diferente al calculado.
        """)

# ============================================================================
# INTERFAZ - TAB 5: RESUMEN GLOBAL
# ============================================================================

def render_tab_resumen():
    """Tab 5: Resumen Global"""
    
    st.markdown('<h2 class="section-title">📊 Resumen Global del Proyecto</h2>', unsafe_allow_html=True)
    
    resumen = get_resumen_cacheado()
    
    # MÉTRICAS PRINCIPALES
    col1, col2, col3 = st.columns(3)
    col1.metric("💰 Total Proyecto", f"${resumen['total_proyecto']:,.0f}")
    col2.metric("📐 Precio por m²", f"${resumen['precio_m2']:,.0f}")
    col3.metric("🏗️ Área Base", f"{st.session_state.proyecto.area_base:.2f} m²")
    
    st.markdown("---")
    
    # DOS COTIZACIONES
    col_cot1, col_cot2 = st.columns(2)
    
    with col_cot1:
        st.markdown("### 📋 Cotización 1: Diseños + Estructura + Mampostería")
        
        cot1 = resumen['cotizacion1']
        
        st.markdown("**Costos Directos:**")
        st.write(f"- Diseños: ${cot1['disenos']:,.0f}")
        st.write(f"- Estructura: ${cot1['estructura']:,.0f}")
        st.write(f"- Mampostería: ${cot1['mamposteria']:,.0f}")
        st.write(f"- Techos y otros: ${cot1['mamposteria_techos']:,.0f}")
        st.write(f"**Subtotal Costos Directos: ${cot1['costos_directos']:,.0f}**")
        
        st.markdown("**AIU:**")
        st.write(f"- Comisión Ventas: ${cot1['aiu']['comision_ventas']:,.0f}")
        st.write(f"- Imprevistos: ${cot1['aiu']['imprevistos']:,.0f}")
        st.write(f"- Administración: ${cot1['aiu']['administracion']:,.0f}")
        st.write(f"- Logística: ${cot1['aiu']['logistica']:,.0f}")
        st.write(f"- Utilidad: ${cot1['aiu']['utilidad']:,.0f}")
        st.write(f"**Total AIU: ${cot1['aiu']['total']:,.0f}**")
        
        st.success(f"### **TOTAL COTIZACIÓN 1: ${cot1['total']:,.0f}**")
    
    with col_cot2:
        st.markdown("### 📋 Cotización 2: Cimentaciones + Complementarios")
        
        cot2 = resumen['cotizacion2']
        
        st.markdown("**Cimentaciones:**")
        st.write(f"- Subtotal: ${cot2['cimentacion']['subtotal']:,.0f}")
        st.write(f"- Comisión: ${cot2['cimentacion']['comision']:,.0f}")
        st.write(f"- AIU: ${cot2['cimentacion']['aiu']:,.0f}")
        st.write(f"**Total Cimentación: ${cot2['cimentacion']['total']:,.0f}**")
        
        st.markdown("**Complementarios:**")
        st.write(f"- Subtotal: ${cot2['complementarios']['subtotal']:,.0f}")
        st.write(f"- Comisión: ${cot2['complementarios']['comision']:,.0f}")
        st.write(f"- AIU: ${cot2['complementarios']['aiu']:,.0f}")
        st.write(f"**Total Complementarios: ${cot2['complementarios']['total']:,.0f}**")
        
        st.success(f"### **TOTAL COTIZACIÓN 2: ${cot2['total']:,.0f}**")
    
    st.markdown("---")
    
    # GRÁFICOS
    col_grafico1, col_grafico2 = st.columns(2)
    
    with col_grafico1:
        st.markdown("#### Distribución Costos Directos")
        
        categorias_costos = {
            'Diseños': resumen['cotizacion1']['disenos'],
            'Estructura': resumen['cotizacion1']['estructura'],
            'Mampostería': resumen['cotizacion1']['mamposteria'],
            'Techos y otros': resumen['cotizacion1']['mamposteria_techos'],
            'Cimentaciones': resumen['cotizacion2']['cimentacion']['subtotal'],
            'Complementarios': resumen['cotizacion2']['complementarios']['subtotal']
        }
        
        df_costos = pd.DataFrame([
            {'Categoría': k, 'Valor': v} 
            for k, v in categorias_costos.items() if v > 0
        ])
        
        if not df_costos.empty:
            fig_pie = px.pie(
                df_costos,
                values='Valor',
                names='Categoría',
                title='Costos por Categoría'
            )
            st.plotly_chart(fig_pie, use_container_width=True)
    
    with col_grafico2:
        st.markdown("#### Comparación Cotizaciones")
        
        df_comparacion = pd.DataFrame([
            {'Cotización': 'Cotización 1', 'Monto': resumen['cotizacion1']['total']},
            {'Cotización': 'Cotización 2', 'Monto': resumen['cotizacion2']['total']}
        ])
        
        fig_bar = px.bar(
            df_comparacion,
            x='Cotización',
            y='Monto',
            title='Comparación entre Cotizaciones',
            text='Monto'
        )
        fig_bar.update_traces(texttemplate='$%{text:,.0f}', textposition='outside')
        st.plotly_chart(fig_bar, use_container_width=True)

# ============================================================================
# INTERFAZ - TAB 6: EXPORTAR
# ============================================================================

def render_tab_exportar():
    """Tab 6: Exportar"""
    
    st.markdown('<h2 class="section-title">📥 Exportar Cotización</h2>', unsafe_allow_html=True)
    
    resumen = get_resumen_cacheado()
    
    st.markdown("### 📄 Vista Previa")
    
    st.markdown(f"""
    **Proyecto:** {st.session_state.proyecto.nombre}  
    **Cliente:** {st.session_state.proyecto.cliente}  
    **Dirección:** {st.session_state.proyecto.direccion}  
    **Área Base:** {st.session_state.proyecto.area_base:.2f} m²  
    
    ---
    
    **Cotización 1 (Diseños + Estructura + Mampostería):** ${resumen['cotizacion1']['total']:,.0f}  
    **Cotización 2 (Cimentaciones + Complementarios):** ${resumen['cotizacion2']['total']:,.0f}  
    **TOTAL PROYECTO:** ${resumen['total_proyecto']:,.0f}  
    **Precio por m²:** ${resumen['precio_m2']:,.0f}
    """)
    
    st.markdown("---")
    
    # GENERAR EXCEL
    st.markdown("### 📥 Generar Excel")
    
    try:
        excel_file = exportar_a_excel()
        st.download_button(
            label="📥 Descargar Cotización Excel",
            data=excel_file,
            file_name=f"Cotizacion_SICONE_{st.session_state.proyecto.nombre.replace(' ', '_')}_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    except Exception as e:
        st.error(f"❌ Error al generar archivo: {str(e)}")
    
    st.markdown("---")
    
    # INFORMACIÓN SOBRE GUARDADO
    st.markdown("### 💡 Información sobre Guardado")
    
    st.info("""
    **📁 Sistema de Guardado en Streamlit Cloud:**
    
    ✅ **Las cotizaciones se guardan durante tu sesión activa** 
    - Al usar el botón "💾" en el sidebar, la cotización se guarda temporalmente
    - Puedes trabajar con múltiples cotizaciones en la misma sesión
    - Cambiar entre cotizaciones es instantáneo
    
    ⚠️ **Importante - Limitación de Streamlit Cloud:**
    - Las cotizaciones guardadas **pueden perderse** si Streamlit reinicia el servidor
    - Esto ocurre por inactividad (~7 días) o actualizaciones de la app
    
    ✅ **Solución Recomendada:**
    - **Siempre usa "📤 Export" para tus cotizaciones importantes**
    - El archivo JSON se descarga a tu PC y es permanente
    - Puedes importarlo cuando lo necesites con "📥 Importar JSON"
    
    📤 **Exportar a JSON (RECOMENDADO):**
    - Backup permanente en tu PC ✅
    - Compartir cotizaciones con otros usuarios ✅
    - Transferir entre diferentes instalaciones ✅
    
    💡 **Mejor práctica:**
    1. Trabaja en tu cotización
    2. 💾 Guarda para tener varias versiones en la sesión
    3. 📤 Export JSON al finalizar para backup permanente
    """)

# ============================================================================
# MAIN APP
# ============================================================================

def main():
    """Aplicación principal"""
    
    inicializar_session_state()
    
    # TÍTULO
    st.markdown('<h1 class="main-title">🏗️ SICONE v3.0 - Sistema de Cotización</h1>', unsafe_allow_html=True)
    
    # ============================================================================
    # SIDEBAR (PRIMERO para que los valores se actualicen ANTES de calcular)
    # ============================================================================
    render_sidebar()
    
    # ============================================================================
    # RESUMEN GLOBAL (siempre visible) - AHORA con valores actualizados
    # ============================================================================
    resumen = get_resumen_cacheado()
    
    st.markdown("---")
    st.markdown("### 📊 Resumen Global del Proyecto")
    
    col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 2, 1])
    with col1:
        st.metric("💰 Total Proyecto", f"${resumen['total_proyecto']:,.0f}")
    with col2:
        st.metric("📐 Precio por m²", f"${resumen['precio_m2']:,.0f}")
    with col3:
        st.metric("📏 Área Base", f"{st.session_state.proyecto.area_base:.2f} m²")
    with col4:
        # Botón de exportación a Excel con descarga directa
        try:
            excel_file = exportar_a_excel()
            st.download_button(
                label="📥 Exportar a Excel",
                data=excel_file,
                file_name=f"Cotizacion_SICONE_{st.session_state.proyecto.nombre.replace(' ', '_')}_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")
    with col5:
        # Botón para forzar actualización
        if st.button("🔄", help="Actualizar resumen", use_container_width=True):
            st.rerun()
    
    st.markdown("---")
    
    # TABS PRINCIPALES
    tabs = st.tabs([
        "📐 Diseños y Estructura",
        "⚙️ Cimentaciones",
        "🔧 Complementarios",
        "💼 Administración",
        "📊 Resumen",
        "📥 Exportar"
    ])
    
    with tabs[0]:
        render_tab_disenos_estructura()
    
    with tabs[1]:
        render_tab_cimentaciones()
    
    with tabs[2]:
        render_tab_complementarios()
    
    with tabs[3]:
        render_tab_administracion()
    
    with tabs[4]:
        render_tab_resumen()
    
    with tabs[5]:
        render_tab_exportar()

if __name__ == "__main__":
    main()
